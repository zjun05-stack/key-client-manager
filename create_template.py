"""Generate the Excel template file. Called by build workflow."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "客户名单"

hf = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ha = Alignment(horizontal="center", vertical="center")
tb = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))

ws1["A1"], ws1["B1"] = "序号", "客户名称"
ws1.column_dimensions["A"].width = 8
ws1.column_dimensions["B"].width = 35
for col in ["A", "B"]:
    c = ws1[f"{col}1"]
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

bf = Font(name="Microsoft YaHei", size=11)
ba = Alignment(horizontal="center", vertical="center")
na = Alignment(horizontal="left", vertical="center")

for i in range(1, 21):
    r = i + 1
    ws1[f"A{r}"] = i
    ws1[f"A{r}"].font = bf; ws1[f"A{r}"].alignment = ba; ws1[f"A{r}"].border = tb
    ws1[f"B{r}"].font = bf; ws1[f"B{r}"].alignment = na; ws1[f"B{r}"].border = tb
    ws1.row_dimensions[r].height = 22
ws1.freeze_panes = "A2"

ws2 = wb.create_sheet("联系记录")
for col_letter, title, width in [("A", "序号", 8), ("B", "客户名称", 35),
                                  ("C", "联系方式", 14), ("D", "联系日期", 14),
                                  ("E", "备注", 40)]:
    c = ws2[f"{col_letter}1"]
    c.value = title; c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    ws2.column_dimensions[col_letter].width = width

dv = DataValidation(type="list", formula1='"电话,拜访,微信,邮件,其他"', allow_blank=True)
ws2.add_data_validation(dv)
dv.add("C2:C1000")
for row in range(2, 1002):
    ws2[f"D{row}"].number_format = "YYYY-MM-DD"
ws2.freeze_panes = "A2"

wb.save("clients.xlsx")
print("clients.xlsx created")
