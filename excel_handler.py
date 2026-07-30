"""Excel 读写层：客户名单 + 联系记录管理."""

import os
import sys
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Detect app directory (works both as .py and as PyInstaller .exe)
if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_FILE = os.path.join(APP_DIR, "clients.xlsx")

BODY_FONT = Font(name="Microsoft YaHei", size=11)
BODY_ALIGN = Alignment(horizontal="center", vertical="center")
NAME_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Cache: avoid re-reading Excel constantly
_cache = {"customers": None, "records": None, "records_week": None, "week_key": None}


def _current_week_key():
    """Return (year, week_number) for cache invalidation."""
    today = date.today()
    return (today.isocalendar()[0], today.isocalendar()[1])


def _invalidate_cache():
    """Clear all cached data."""
    _cache["customers"] = None
    _cache["records"] = None
    _cache["records_week"] = None


def _safe_load_workbook(read_only=True):
    """Load workbook with error handling for locked files."""
    try:
        return load_workbook(EXCEL_FILE, data_only=read_only, read_only=read_only)
    except PermissionError:
        raise PermissionError(
            "Excel 文件正在被其他程序占用（可能正用 Microsoft Excel 打开）。\n"
            "请关闭 Excel 后重试。"
        )


def _get_week_range(today=None):
    """Return (monday, sunday) for the current week."""
    if today is None:
        today = date.today()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def load_customers():
    """Read customer list from Sheet 1. Returns list of {index, name}."""
    if _cache["customers"] is not None:
        return _cache["customers"]
    wb = _safe_load_workbook()
    try:
        ws = wb["客户名单"]
        customers = []
        for row in ws.iter_rows(min_row=2, max_row=21, min_col=1, max_col=2, values_only=True):
            idx, name = row
            if name and str(name).strip():
                customers.append({"index": idx, "name": str(name).strip()})
        _cache["customers"] = customers
        return customers
    except KeyError:
        raise KeyError("Excel 文件中缺少「客户名单」工作表，请检查模板是否正确。")
    finally:
        wb.close()


def get_contact_records():
    """Read all contact records from Sheet 2. Returns list of dicts."""
    if _cache["records"] is not None:
        return _cache["records"]
    wb = _safe_load_workbook()
    try:
        ws = wb["联系记录"]
        records = []
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=5, values_only=True):
            idx, name, method, record_date, notes = row
            if name and str(name).strip():
                records.append({
                    "index": idx,
                    "name": str(name).strip(),
                    "method": str(method).strip() if method else "",
                    "date": _parse_date(record_date),
                    "notes": str(notes).strip() if notes else "",
                })
        _cache["records"] = records
        return records
    except KeyError:
        raise KeyError("Excel 文件中缺少「联系记录」工作表，请检查模板是否正确。")
    finally:
        wb.close()


def _parse_date(value):
    """Parse a date value from openpyxl (could be datetime, date, or string)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"]:
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def get_this_week_contacted():
    """Return set of customer names contacted this week (Mon-Sun)."""
    week_key = _current_week_key()
    if _cache["records_week"] is not None and _cache["week_key"] == week_key:
        return _cache["records_week"]

    monday, sunday = _get_week_range()
    records = get_contact_records()
    contacted = set()
    for r in records:
        if r["date"] and monday <= r["date"] <= sunday:
            contacted.add(r["name"])

    _cache["records_week"] = contacted
    _cache["week_key"] = week_key
    return contacted


def get_last_contact_date_map():
    """Return {name: last_contact_date} mapping for all customers (one pass)."""
    records = get_contact_records()
    date_map = {}
    for r in records:
        if r["date"]:
            name = r["name"]
            if name not in date_map or r["date"] > date_map[name]:
                date_map[name] = r["date"]
    return date_map


def add_contact_record(name, method, record_date, notes=""):
    """Append a contact record to Sheet 2."""
    wb = _safe_load_workbook(read_only=False)
    try:
        ws = wb["联系记录"]
        next_row = ws.max_row + 1
        if next_row < 2:
            next_row = 2

        ws[f"A{next_row}"] = next_row - 1
        ws[f"B{next_row}"] = name
        ws[f"C{next_row}"] = method
        ws[f"D{next_row}"] = record_date
        ws[f"D{next_row}"].number_format = "YYYY-MM-DD"
        ws[f"E{next_row}"] = notes

        for col in ["A", "B", "C", "D", "E"]:
            cell = ws[f"{col}{next_row}"]
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col == "B":
                cell.alignment = NAME_ALIGN
            else:
                cell.alignment = BODY_ALIGN

        wb.save(EXCEL_FILE)
        _invalidate_cache()
    finally:
        wb.close()


def get_week_info():
    """Return (week_number, monday, friday) for display."""
    today = date.today()
    week_num = today.isocalendar()[1]
    monday, _ = _get_week_range(today)
    friday = monday + timedelta(days=4)
    return week_num, monday, friday
